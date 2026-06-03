"""
Teacher lesson source file uploads — extract text from PDF/DOCX/TXT for AI generation.
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
import tempfile
import uuid
from datetime import datetime, timezone

from teacher_teaching_pages import MAX_SOURCE_TEXT

ALLOWED_SOURCE_EXTENSIONS = frozenset({"pdf", "docx", "txt", "ppt", "pptx", "xlsx"})
OFFICE_TO_PDF_EXTS = frozenset({"ppt", "pptx"})
MAX_SOURCE_FILE_BYTES = 10 * 1024 * 1024
MAX_SOURCE_FILES_PER_TEACHER = 12
PREVIEW_CHARS = 600


def _now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def allowed_source_extension(filename: str) -> bool:
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[-1].lower()
    return ext in ALLOWED_SOURCE_EXTENSIONS


def extract_text_from_bytes(data: bytes, ext: str) -> str:
    ext = str(ext or "").lower().strip()
    if ext == "txt":
        for enc in ("utf-8", "utf-8-sig", "latin-1"):
            try:
                return data.decode(enc)
            except UnicodeDecodeError:
                continue
        return data.decode("utf-8", errors="replace")

    if ext == "pdf":
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("PDF support is not installed on the server") from exc
        import io

        reader = PdfReader(io.BytesIO(data))
        parts = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
        return "\n".join(parts)

    if ext == "docx":
        try:
            from docx import Document
        except ImportError as exc:
            raise RuntimeError("Word support is not installed on the server") from exc
        import io

        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs if p.text)

    if ext in OFFICE_TO_PDF_EXTS:
        return _extract_text_via_office_pdf(data, ext)

    if ext == "xlsx":
        return _extract_text_from_xlsx(data)

    raise ValueError(f"Unsupported file type: {ext}")


def _find_soffice_binary() -> str | None:
    for name in ("soffice", "libreoffice"):
        found = shutil.which(name)
        if found:
            return found
    for path in (
        "/usr/bin/soffice",
        "/usr/bin/libreoffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ):
        if os.path.isfile(path) and os.access(path, os.X_OK):
            return path
    return None


def _extract_text_via_office_pdf(data: bytes, ext: str) -> str:
    soffice = _find_soffice_binary()
    if not soffice:
        raise RuntimeError(
            "PowerPoint support requires LibreOffice (soffice) on the server"
        )
    with tempfile.TemporaryDirectory() as td:
        src = os.path.join(td, f"source.{ext}")
        out_dir = os.path.join(td, "out")
        os.makedirs(out_dir, exist_ok=True)
        with open(src, "wb") as fh:
            fh.write(data)
        profile_dir = os.path.join(td, "lo_profile")
        os.makedirs(profile_dir, exist_ok=True)
        cmd = [
            soffice,
            "--headless",
            "--norestore",
            "--invisible",
            f"-env:UserInstallation=file://{profile_dir}",
            "--convert-to",
            "pdf",
            "--outdir",
            out_dir,
            src,
        ]
        proc = subprocess.run(cmd, capture_output=True, timeout=120, check=False)
        if proc.returncode != 0:
            err = (proc.stderr or proc.stdout or b"").decode("utf-8", errors="replace")[:300]
            raise RuntimeError(f"Could not convert {ext} to PDF: {err}")
        pdf_path = os.path.join(out_dir, "source.pdf")
        if not os.path.isfile(pdf_path):
            for name in os.listdir(out_dir):
                if name.lower().endswith(".pdf"):
                    pdf_path = os.path.join(out_dir, name)
                    break
        if not os.path.isfile(pdf_path):
            raise RuntimeError("PDF conversion produced no output file")
        with open(pdf_path, "rb") as fh:
            return extract_text_from_bytes(fh.read(), "pdf")


def _extract_text_from_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel support requires openpyxl on the server") from exc
    import io

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames[:6]:
        ws = wb[sheet_name]
        parts.append(f"=== Sheet: {sheet_name} ===")
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if row_count >= 80:
                parts.append("… (truncated)")
                break
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                parts.append("\t".join(cells))
                row_count += 1
    wb.close()
    return "\n".join(parts)


def normalize_extracted_text(text: str) -> str:
    cleaned = re.sub(r"\r\n?", "\n", str(text or ""))
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def preview_text(text: str, limit: int = PREVIEW_CHARS) -> str:
    t = normalize_extracted_text(text)
    if len(t) <= limit:
        return t
    return t[:limit].rstrip() + "…"


def row_to_dict(row) -> dict:
    text = row["extracted_text"] or ""
    return {
        "id": row["id"],
        "original_name": row["original_name"] or "",
        "status": row["status"] or "staged",
        "char_count": row["char_count"] or len(text),
        "preview": preview_text(text),
        "created_at": row["created_at"] or "",
        "confirmed_at": row["confirmed_at"] or "",
    }


def row_to_detail(row) -> dict:
    payload = row_to_dict(row)
    payload["extracted_text"] = row["extracted_text"] or ""
    return payload


def merge_source_text(paste_text: str, file_texts: list[str], max_total: int = MAX_SOURCE_TEXT) -> str:
    parts = []
    paste = normalize_extracted_text(paste_text)
    if paste:
        parts.append(paste)
    for block in file_texts:
        t = normalize_extracted_text(block)
        if t:
            parts.append(t)
    merged = "\n\n---\n\n".join(parts).strip()
    if len(merged) > max_total:
        merged = merged[:max_total]
    return merged


def teaching_source_upload_dir(base_upload_dir: str) -> str:
    path = os.path.join(base_upload_dir, "teaching-sources")
    os.makedirs(path, exist_ok=True)
    return path


def save_source_file(upload_dir: str, original_name: str, data: bytes) -> tuple[str, str]:
    ext = original_name.rsplit(".", 1)[-1].lower()
    stored = f"{uuid.uuid4().hex}.{ext}"
    dest = os.path.join(upload_dir, stored)
    with open(dest, "wb") as fh:
        fh.write(data)
    return stored, dest


def delete_stored_file(upload_dir: str, stored_name: str | None) -> None:
    if not stored_name:
        return
    base = os.path.basename(stored_name)
    path = os.path.join(upload_dir, base)
    if os.path.isfile(path):
        try:
            os.remove(path)
        except OSError:
            pass
